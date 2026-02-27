"""
Excel writer abstraction: openpyxl baseline; pluggable MCP adapter later.
Writes into pre-defined anchors/ranges from TemplateMap. Does not rename sheets or break layout.
Per T011; FR-007.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Optional

from src.models.template_map import SheetMap, TemplateMap

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Baseline writer using openpyxl. Loads template, writes cells, saves to output path."""

    def __init__(self, template_path: str | Path, template_map: TemplateMap):
        self.template_path = Path(template_path)
        self.template_map = template_map
        self._wb = None

    def load(self) -> None:
        """Load workbook from template."""
        import openpyxl
        self._wb = openpyxl.load_workbook(self.template_path)
        logger.info("Loaded template %s with sheets: %s", self.template_path, self._wb.sheetnames)

    def ensure_loaded(self) -> None:
        if self._wb is None:
            self.load()

    def write_cell(self, sheet_name: str, cell_ref: str, value: Any) -> bool:
        """Write a single cell. cell_ref e.g. 'C5'. Returns True if written."""
        self.ensure_loaded()
        if sheet_name not in self._wb.sheetnames:
            logger.warning("Sheet %s not in workbook", sheet_name)
            return False
        try:
            sheet = self._wb[sheet_name]
            sheet[cell_ref] = value
            return True
        except Exception as e:
            logger.warning("write_cell %s %s failed: %s", sheet_name, cell_ref, e)
            return False

    def write_evidence_columns(
        self,
        sheet_name: str,
        row: int,
        source: str = "",
        evidence: str = "",
        confidence: str = "",
        sheet_map: Optional[SheetMap] = None,
    ) -> None:
        """Write source/evidence/confidence into evidence columns at row. Column letters from SheetMap or default."""
        self.ensure_loaded()
        if sheet_name not in self._wb.sheetnames:
            return
        sheet = self._wb[sheet_name]
        evidence_cols = (sheet_map and sheet_map.evidence_cols) or ["Source", "Evidence", "Confidence"]
        col_letters = evidence_cols if evidence_cols and isinstance(evidence_cols[0], str) and len(evidence_cols[0]) <= 2 else ["G", "H", "I"]
        values = [source, evidence, confidence]
        for i, col in enumerate(col_letters[:3]):
            if i < len(values) and col:
                try:
                    sheet[f"{col}{row}"] = values[i]
                except Exception as e:
                    logger.debug("write_evidence_columns %s: %s", col, e)

    def _col_letter(self, index: int) -> str:
        """1-based column index to letter(s)."""
        s = ""
        i = index
        while i > 0:
            i, r = divmod(i - 1, 26)
            s = chr(65 + r) + s
        return s or "A"

    def write_table(
        self,
        sheet_name: str,
        start_row: int,
        grid: list[list[Any]],
        sheet_map: Optional[SheetMap] = None,
    ) -> int:
        """
        Write a 2D grid into the sheet starting at start_row.
        Uses line_item_column and year_columns for data; never writes into evidence_cols (G, H, I).
        Returns the last row written (1-based).
        """
        self.ensure_loaded()
        if sheet_name not in self._wb.sheetnames:
            logger.warning("Sheet %s not in workbook", sheet_name)
            return start_row
        sheet = self._wb[sheet_name]
        line_col = (sheet_map and sheet_map.line_item_column) or "A"
        year_cols = (sheet_map and sheet_map.year_columns) or []
        evidence_cols = set((sheet_map and sheet_map.evidence_cols) or [])
        # Build data column list: line_item + year_cols, then next letters that are NOT evidence
        col_letters: list[str] = [line_col]
        for c in year_cols:
            if c and c not in col_letters:
                col_letters.append(c)
        max_cols = max(len(row) for row in grid) if grid else 0
        idx = len(col_letters) + 1
        while len(col_letters) < max_cols:
            next_col = self._col_letter(idx)
            if next_col not in evidence_cols:
                col_letters.append(next_col)
            idx += 1
        for r_idx, row in enumerate(grid):
            row_num = start_row + r_idx
            for c_idx, val in enumerate(row):
                if c_idx >= len(col_letters):
                    break
                col = col_letters[c_idx]
                try:
                    sheet[f"{col}{row_num}"] = val
                except Exception as e:
                    logger.debug("write_table cell %s%s: %s", col, row_num, e)
        return start_row + len(grid) - 1 if grid else start_row

    def write_evidence_for_table(
        self,
        sheet_name: str,
        row: int,
        page_ref: str,
        table_id: str,
        confidence: str = "High",
        sheet_map: Optional[SheetMap] = None,
    ) -> None:
        """Write page reference and table_id into evidence columns at the given row."""
        self.write_evidence_columns(
            sheet_name=sheet_name,
            row=row,
            source=page_ref,
            evidence=table_id or "(table)",
            confidence=confidence,
            sheet_map=sheet_map,
        )

    def write_section_header(self, sheet_name: str, row: int, title: str, page_ref: str = "") -> None:
        """Write a section header (e.g. note name) and optional page ref at row."""
        self.ensure_loaded()
        if sheet_name not in self._wb.sheetnames:
            return
        sheet = self._wb[sheet_name]
        try:
            sheet[f"A{row}"] = title
            if page_ref:
                sheet[f"B{row}"] = page_ref
        except Exception as e:
            logger.debug("write_section_header: %s", e)

    def write_interpretation_bullets(
        self,
        sheet_name: str,
        start_row: int,
        section_title: str,
        bullets: list[tuple[str, str]],
        text_col: str = "A",
        ref_col: str = "B",
    ) -> int:
        """Write interpretation bullets: (text, page_ref) per row. Returns last row written."""
        self.ensure_loaded()
        if sheet_name not in self._wb.sheetnames:
            return start_row
        sheet = self._wb[sheet_name]
        row = start_row
        try:
            sheet[f"{text_col}{row}"] = section_title
            row += 1
            for text, ref in bullets:
                sheet[f"{text_col}{row}"] = (text or "")[:2000]
                sheet[f"{ref_col}{row}"] = ref or ""
                row += 1
        except Exception as e:
            logger.debug("write_interpretation_bullets: %s", e)
        return row

    def save(self, output_path: str | Path) -> None:
        """Save workbook to output path."""
        self.ensure_loaded()
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(out)
        logger.info("Saved workbook to %s", out)
