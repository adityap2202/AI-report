"""
Analyze an Excel workbook and generate a TemplateMap JSON (e.g. knowledge_marine_v1.json).
Run before extraction so template_map is populated from the actual workbook.
"""
from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

# Patterns for header detection (year columns)
YEAR_HEADER_PATTERNS = re.compile(r"FY|F\.Y|20\d{2}|\d{4}[-/]\d{2}|year|₹|Rs\.?", re.I)
# Column names that indicate evidence columns
EVIDENCE_COLUMN_NAMES = {"source", "evidence", "confidence", "page", "reference"}
# Sheet name -> type and statement/topic hints
SHEET_TYPE_HINTS = {
    "cf bs": ("financial_statement", "balance_sheet", None),
    "balance sheet": ("financial_statement", "balance_sheet", None),
    "cf is": ("financial_statement", "income_statement", None),
    "income statement": ("financial_statement", "income_statement", None),
    "cf cfs": ("financial_statement", "cash_flow", None),
    "cash flow": ("financial_statement", "cash_flow", None),
    "borrowings": ("note_rollup", None, "borrowings"),
    "ppe": ("note_rollup", None, "ppe"),
    "revenue": ("note_rollup", None, "revenue"),
    "revenue & oi": ("note_rollup", None, "revenue"),
    "expenses": ("note_rollup", None, "expenses"),
    "working capital": ("note_rollup", None, "working_capital"),
    "rpt": ("note_rollup", None, "rpt"),
    "segment": ("note_rollup", None, "segment"),
}


def _col_letter(index: int) -> str:
    """1-based column index to letter(s): 1->A, 27->AA."""
    s = ""
    while index > 0:
        index, r = divmod(index - 1, 26)
        s = chr(65 + r) + s
    return s or "A"


def _detect_header_row(sheet) -> int | None:
    """Scan first 20 rows for a row that looks like year headers. Return 1-based row number."""
    for row_idx in range(1, min(21, sheet.max_row + 1)):
        row_vals = []
        for col_idx in range(1, min(sheet.max_column + 1, 30)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                row_vals.append((col_idx, str(val).strip()))
        if not row_vals:
            continue
        # Check if any cell looks like year/FY
        for col_idx, val in row_vals:
            if YEAR_HEADER_PATTERNS.search(val):
                return row_idx
    return None


def _detect_evidence_columns(sheet, header_row: int) -> list[str]:
    """Find columns whose header is Source, Evidence, Confidence, etc. Return column letters."""
    letters = []
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        val = cell.value
        if val is not None and str(val).strip().lower() in EVIDENCE_COLUMN_NAMES:
            letters.append(_col_letter(col_idx))
    return letters


def _detect_year_columns(sheet, header_row: int) -> list[str]:
    """Columns in header row that match year/FY pattern."""
    letters = []
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        val = cell.value
        if val is not None and YEAR_HEADER_PATTERNS.search(str(val)):
            letters.append(_col_letter(col_idx))
    return letters


def _detect_line_item_column(sheet, header_row: int, year_columns: list[str]) -> str:
    """Typically column A or B. Prefer column that has text in data rows and is not a year col."""
    year_col_indices = {ord(c) - 64 if len(c) == 1 else 0 for c in year_columns}
    # Simple: A=1, B=2; use A if not in year cols else B
    for col_idx in (1, 2):
        if col_idx not in year_col_indices:
            return _col_letter(col_idx)
    return "A"


def analyze_workbook(workbook_path: str | Path) -> dict[str, Any]:
    """
    Open workbook with openpyxl, analyze each sheet, return a dict suitable for TemplateMap JSON.
    """
    import openpyxl
    path = Path(workbook_path)
    if not path.is_file():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    workbook_name = path.name
    sheets_out: dict[str, Any] = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        header_row = _detect_header_row(sheet)
        year_cols = _detect_year_columns(sheet, header_row) if header_row else []
        line_item_col = _detect_line_item_column(sheet, header_row or 1, year_cols) if header_row else "A"
        evidence_cols = _detect_evidence_columns(sheet, header_row or 1) if header_row else []
        if not evidence_cols:
            evidence_cols = ["Source", "Evidence", "Confidence"]  # placeholder names
        key = sheet_name.strip().lower()
        type_hint = SHEET_TYPE_HINTS.get(key)
        if type_hint:
            layout_type, statement, topic = type_hint
            sheet_def: dict[str, Any] = {
                "type": layout_type,
                "header_row": header_row,
                "line_item_column": line_item_col,
                "year_columns": year_cols,
                "evidence_cols": evidence_cols,
            }
            if statement:
                sheet_def["statement"] = statement
            if topic:
                sheet_def["topic"] = topic
        else:
            sheet_def = {
                "type": "note_rollup",
                "header_row": header_row,
                "line_item_column": line_item_col,
                "year_columns": year_cols,
                "evidence_cols": evidence_cols,
            }
        sheets_out[sheet_name] = sheet_def
    wb.close()
    return {
        "workbook_name": workbook_name,
        "sheets": sheets_out,
        "write_rules": {"skip_formula_cells": True, "skip_non_empty_cells": False},
        "sheet_detection_rules": {
            "year_header_patterns": ["FY", "F.Y", "20", "202"],
            "line_item_column_candidates": ["A", "B"],
            "evidence_column_names": list(EVIDENCE_COLUMN_NAMES),
        },
    }


def write_template_map_from_workbook(
    workbook_path: str | Path,
    output_path: str | Path,
) -> Path:
    """Analyze workbook and write TemplateMap JSON to output_path (e.g. template_map/knowledge_marine_v1.json)."""
    data = analyze_workbook(workbook_path)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    logger.info("Wrote TemplateMap to %s from %s", out, workbook_path)
    return out
